from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from os.path import exists

FILE_NAME = "attendance.xlsx"


def save_to_excel(date, subject, roll, name, status):

    if exists(FILE_NAME):
        wb = load_workbook(FILE_NAME)
        ws = wb.active

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        headers = [
            "Date",
            "Subject",
            "Roll No",
            "Student Name",
            "Status"
        ]

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        for col, head in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=col)

            cell.value = head
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

    ws.append([
        date,
        subject,
        roll,
        name,
        status
    ])

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value))
            if cell.value
            else 0
            for cell in column_cells
        )

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = length + 5

    wb.save(FILE_NAME)